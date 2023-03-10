from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference

def countOccurence(a, b):
    wordsA = []
    wordsB = []
    ### split sentences into words
    for line in a:
        wordsA.extend(line.split(' '))
    for line in b:
        wordsB.extend(line.split(' '))
    
    ### define 26 letters. All characters not in this list will be removed.
    alphabets = [chr(i) for i in range(97, 123)]
    
    ### count word occurance. At same time remove punctuations
    countA = {}
    countB = {}
    for word in wordsA:
        ### consider below scenario:
            ### we will run A / B test.
        if len(word) == 1 and word not in alphabets:
            continue
        if word[0] not in alphabets:
            word = word[1:]
        if word[-1] not in alphabets:
            word = word[:-1]
        if len(word) == 0:
            continue
        if word not in countA:
            countA[word] = 0
        if word not in countB:
            countB[word] = 0
        countA[word] += 1
    for word in wordsB:
        ### consider below scenario:
            ### we will run A / B test.
        if len(word) == 1 and word not in alphabets:
            continue
        if word[0] not in alphabets:
            word = word[1:]
        if word[-1] not in alphabets:
            word = word[:-1]
        if len(word) == 0:
            continue
        if word not in countB:
            countB[word] = 0
        countB[word] += 1

    
    res = []
    for key, val in countA.items():
        if val > countB[key]:
            res.append([key, val, countB[key]])
    res.sort(key = lambda x: x[1], reverse=True)
    
    output = './WordOccurence.xlsx'
    wb = Workbook()
    ws = wb.active
    ws['A1'], ws['B1'], ws['C1'] = 'Word', '#(Occurrences) in A', '#(Occurrences) in B'

    for i in range(2, len(res) + 2):
        ws[f'A{i}'], ws[f'B{i}'], ws[f'C{i}'] = res[i - 2][0], res[i - 2][1], res[i - 2][2]
    
    
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Top 5 Word Occurence in fileA"
    chart1.y_axis.title = 'Num. of Occurence'
    chart1.x_axis.title = 'Words'

    data = Reference(ws, min_col=2, min_row=1, max_row=6, max_col=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    ws.add_chart(chart1, "G2")



    wb.save(output)












if __name__ == '__main__':
    A = open('./fileA.txt', 'rt')
    B = open('./fileB.txt', 'rt')

    a = [line.strip().lower() for line in A.readlines()]
    b = [line.strip().lower() for line in B.readlines()]

    A.close()
    B.close()

    countOccurence(a, b)



    

    
